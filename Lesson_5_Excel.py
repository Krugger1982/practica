from openpyxl import Workbook, load_workbook
import random

dest_filename = 'New_book.xlsx'
wb = load_workbook(filename = dest_filename)
ws2 = wb.active
for columns in range(1, 13):
    for rows in range (1, 13):
        ws2.cell(column = columns, row = rows, value = random.randint(1, 100))
wb.save(filename = dest_filename)
